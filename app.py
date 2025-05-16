# IMPORTANT: Place this code at the very top of your script
# before importing any other libraries

# 1. Set environment variables before any imports
import os
os.environ['TF_CPP_MIN_LOG_LEVEL'] = '3'  # 3 = ERROR only (most aggressive)
os.environ['MEDIAPIPE_DISABLE_GPU'] = '1'
os.environ['AUTOGRAPH_VERBOSITY'] = '0'  # Reduce TensorFlow Autograph logs
os.environ['TF_ENABLE_ONEDNN_OPTS'] = '0'  # Disable oneDNN optimization messages
os.environ['PYTHONWARNINGS'] = 'ignore'  # Additional warning suppression
os.environ['TF_CPP_MAX_VLOG_LEVEL'] = '0'  # Suppress TensorFlow verbose logging including XNNPACK messages

# 2. Configure Python warnings
import warnings
warnings.filterwarnings('ignore', category=UserWarning)
warnings.filterwarnings('ignore', category=DeprecationWarning)
warnings.filterwarnings('ignore', category=FutureWarning)

# 3. Configure standard logging
import logging
logging.basicConfig(level=logging.ERROR)
for logger_name in ['tensorflow', 'mediapipe', 'absl', 'matplotlib', 'google']:
    logging.getLogger(logger_name).setLevel(logging.ERROR)

# 4. Now you can import the rest of your libraries
from flask import Flask, request, jsonify, render_template, send_file
import cv2
import mediapipe as mp
import numpy as np
import threading
import time
import pandas as pd
from sklearn.linear_model import LinearRegression
from sklearn.preprocessing import PolynomialFeatures
from flask_cors import CORS
import torch
import base64
from io import BytesIO
from PIL import Image
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime

# Initialize Flask application
app = Flask(__name__, static_folder='public', template_folder='public')
CORS(app, resources={
    r"/*": {
        "origins": ["http://localhost:3000", "http://127.0.0.1:3000"],
        "methods": ["GET", "POST", "OPTIONS"],
        "allow_headers": ["Content-Type"],
        "supports_credentials": True
    }
})

# Load bone density dataset
try:
    print("Loading bone density dataset...")
    df = pd.read_csv("Bone Mineral Density.txt", sep='\t')
    print(f"Dataset loaded successfully with {len(df)} records")
    df_male = df[df.gender == 'male']
    df_female = df[df.gender == 'female']
    min_bmd = df.spnbmd.min()
    max_bmd = df.spnbmd.max()
    print(f"Male records: {len(df_male)}, Female records: {len(df_female)}")
    print(f"BMD range: {min_bmd} to {max_bmd}")
except Exception as e:
    print(f"Error loading bone density dataset: {str(e)}")
    df = None
    df_male = None
    df_female = None
    min_bmd = 0
    max_bmd = 1

class ParametricBMDModel:
    def __init__(self, peak_age=30):
        self.peak_age = peak_age

    def fit(self, X, y):
        X = X.flatten()
        self.X_pre = X[X <= self.peak_age].reshape(-1, 1)
        self.y_pre = y[X <= self.peak_age]
        self.X_post = X[X > self.peak_age].reshape(-1, 1)
        self.y_post = y[X > self.peak_age]

        if len(self.X_pre) > 0:
            self.poly_pre = PolynomialFeatures(degree=2)
            X_pre_poly = self.poly_pre.fit_transform(self.X_pre)
            self.model_pre = LinearRegression().fit(X_pre_poly, self.y_pre)
        else:
            self.model_pre = None

        if len(self.X_post) > 0:
            self.poly_post = PolynomialFeatures(degree=1)
            X_post_poly = self.poly_post.fit_transform(self.X_post)
            self.model_post = LinearRegression().fit(X_post_poly, self.y_post)
        else:
            self.model_post = None

    def predict(self, X):
        X = np.array(X).reshape(-1, 1)
        preds = []
        for x in X:
            if x <= self.peak_age and self.model_pre:
                x_poly = self.poly_pre.transform([[x[0]]])
                preds.append(self.model_pre.predict(x_poly)[0])
            elif x > self.peak_age and self.model_post:
                x_poly = self.poly_post.transform([[x[0]]])
                preds.append(self.model_post.predict(x_poly)[0])
            elif self.model_pre:
                x_poly = self.poly_pre.transform([[x[0]]])
                preds.append(self.model_pre.predict(x_poly)[0])
            elif self.model_post:
                x_poly = self.poly_post.transform([[x[0]]])
                preds.append(self.model_post.predict(x_poly)[0])
            else:
                preds.append(np.nan)
        print(preds)
        return np.array(preds)

# Initialize bone density models if dataset is available
if df is not None:
    model_male = ParametricBMDModel()
    model_male.fit(df_male[['age']].values, df_male['spnbmd'].values)

    model_female = ParametricBMDModel()
    model_female.fit(df_female[['age']].values, df_female['spnbmd'].values)
else:
    model_male = None
    model_female = None


# Initialize MediaPipe Pose with specific configuration to avoid warnings
mp_pose = mp.solutions.pose
mp_drawing = mp.solutions.drawing_utils

# Global variables to store workout data
workout_data = {
    "repCount": 0,
    "duration": 0,
    "calories": 0,
    "stage": None
}

# Flag to control the workout monitoring thread
is_workout_running = False
workout_thread = None
start_time = 0

def calculate_angle(a, b, c):
    """Calculate the angle between three points."""
    a = np.array(a)
    b = np.array(b)
    c = np.array(c)
    
    radians = np.arctan2(c[1]-b[1], c[0]-b[0]) - np.arctan2(a[1]-b[1], a[0]-b[0])
    angle = np.abs(radians*180.0/np.pi)
    
    if angle > 180.0:
        angle = 360-angle
        
    return angle

def detect_exercise(landmarks, workout_type):
    """Detect exercise based on pose landmarks and workout type."""
    try:
        if workout_type == "bicep_curl":
            # Get coordinates for both arms
            left_shoulder = [landmarks[mp_pose.PoseLandmark.LEFT_SHOULDER.value].x,
                           landmarks[mp_pose.PoseLandmark.LEFT_SHOULDER.value].y]
            left_elbow = [landmarks[mp_pose.PoseLandmark.LEFT_ELBOW.value].x,
                         landmarks[mp_pose.PoseLandmark.LEFT_ELBOW.value].y]
            left_wrist = [landmarks[mp_pose.PoseLandmark.LEFT_WRIST.value].x,
                         landmarks[mp_pose.PoseLandmark.LEFT_WRIST.value].y]
            
            right_shoulder = [landmarks[mp_pose.PoseLandmark.RIGHT_SHOULDER.value].x,
                            landmarks[mp_pose.PoseLandmark.RIGHT_SHOULDER.value].y]
            right_elbow = [landmarks[mp_pose.PoseLandmark.RIGHT_ELBOW.value].x,
                          landmarks[mp_pose.PoseLandmark.RIGHT_ELBOW.value].y]
            right_wrist = [landmarks[mp_pose.PoseLandmark.RIGHT_WRIST.value].x,
                          landmarks[mp_pose.PoseLandmark.RIGHT_WRIST.value].y]
            
            # Calculate angles for both arms
            left_angle = calculate_angle(left_shoulder, left_elbow, left_wrist)
            right_angle = calculate_angle(right_shoulder, right_elbow, right_wrist)
            
            # Count rep when either arm is curled (angle < 30 degrees)
            # and the other arm is extended (angle > 160 degrees)
            return (left_angle < 30 and right_angle > 160) or (right_angle < 30 and left_angle > 160)
        
        elif workout_type == "squat":
            # Get coordinates for both legs
            left_hip = [landmarks[mp_pose.PoseLandmark.LEFT_HIP.value].x,
                       landmarks[mp_pose.PoseLandmark.LEFT_HIP.value].y]
            left_knee = [landmarks[mp_pose.PoseLandmark.LEFT_KNEE.value].x,
                        landmarks[mp_pose.PoseLandmark.LEFT_KNEE.value].y]
            left_ankle = [landmarks[mp_pose.PoseLandmark.LEFT_ANKLE.value].x,
                         landmarks[mp_pose.PoseLandmark.LEFT_ANKLE.value].y]
            
            right_hip = [landmarks[mp_pose.PoseLandmark.RIGHT_HIP.value].x,
                        landmarks[mp_pose.PoseLandmark.RIGHT_HIP.value].y]
            right_knee = [landmarks[mp_pose.PoseLandmark.RIGHT_KNEE.value].x,
                         landmarks[mp_pose.PoseLandmark.RIGHT_KNEE.value].y]
            right_ankle = [landmarks[mp_pose.PoseLandmark.RIGHT_ANKLE.value].x,
                          landmarks[mp_pose.PoseLandmark.RIGHT_ANKLE.value].y]
            
            # Calculate angles for both legs
            left_angle = calculate_angle(left_hip, left_knee, left_ankle)
            right_angle = calculate_angle(right_hip, right_knee, right_ankle)
            
            # Count rep when both knees are bent (angle < 100 degrees)
            # and hips are below knees (checking y-coordinates)
            return (left_angle < 100 and right_angle < 100 and 
                   left_hip[1] > left_knee[1] and right_hip[1] > right_knee[1])
        
        elif workout_type == "pushup":
            # Get coordinates for both arms
            left_shoulder = [landmarks[mp_pose.PoseLandmark.LEFT_SHOULDER.value].x,
                           landmarks[mp_pose.PoseLandmark.LEFT_SHOULDER.value].y]
            left_elbow = [landmarks[mp_pose.PoseLandmark.LEFT_ELBOW.value].x,
                         landmarks[mp_pose.PoseLandmark.LEFT_ELBOW.value].y]
            left_wrist = [landmarks[mp_pose.PoseLandmark.LEFT_WRIST.value].x,
                         landmarks[mp_pose.PoseLandmark.LEFT_WRIST.value].y]
            
            right_shoulder = [landmarks[mp_pose.PoseLandmark.RIGHT_SHOULDER.value].x,
                            landmarks[mp_pose.PoseLandmark.RIGHT_SHOULDER.value].y]
            right_elbow = [landmarks[mp_pose.PoseLandmark.RIGHT_ELBOW.value].x,
                          landmarks[mp_pose.PoseLandmark.RIGHT_ELBOW.value].y]
            right_wrist = [landmarks[mp_pose.PoseLandmark.RIGHT_WRIST.value].x,
                          landmarks[mp_pose.PoseLandmark.RIGHT_WRIST.value].y]
            
            # Calculate angles for both arms
            left_angle = calculate_angle(left_shoulder, left_elbow, left_wrist)
            right_angle = calculate_angle(right_shoulder, right_elbow, right_wrist)
            
            # Count rep when both arms are bent (angle < 90 degrees)
            # and body is parallel to ground (checking shoulder and hip alignment)
            left_hip = landmarks[mp_pose.PoseLandmark.LEFT_HIP.value].y
            right_hip = landmarks[mp_pose.PoseLandmark.RIGHT_HIP.value].y
            shoulder_hip_alignment = abs(left_shoulder[1] - left_hip) < 0.1 and abs(right_shoulder[1] - right_hip) < 0.1
            
            return (left_angle < 90 and right_angle < 90 and shoulder_hip_alignment)
        
        return False
    except Exception as e:
        print(f"Error in exercise detection: {str(e)}")
        return False

def workout_monitoring(workout_type):
    """Monitor workout using OpenCV and MediaPipe."""
    global workout_data, is_workout_running, start_time
    
    # Initialize video capture
    cap = cv2.VideoCapture(0)
    
    # Ensure camera is opened successfully
    if not cap.isOpened():
        print("Error: Could not open camera.")
        is_workout_running = False
        return
    
    # Set specific camera properties to help with MediaPipe
    cap.set(cv2.CAP_PROP_FRAME_WIDTH, 640)
    cap.set(cv2.CAP_PROP_FRAME_HEIGHT, 480)
    
    # Get frame dimensions
    width = int(cap.get(cv2.CAP_PROP_FRAME_WIDTH))
    height = int(cap.get(cv2.CAP_PROP_FRAME_HEIGHT))
    print(f"Camera resolution: {width}x{height}")
    
    # Initialize pose with static images mode off for video
    with mp_pose.Pose(
        min_detection_confidence=0.7,  # Increased confidence threshold
        min_tracking_confidence=0.7,   # Increased tracking confidence
        model_complexity=1,
        static_image_mode=False,
        enable_segmentation=False,
        smooth_segmentation=False
    ) as pose:
        
        # Initialize variables
        rep_count = 0
        stage = None
        start_time = time.time()
        last_rep_time = time.time()
        rep_cooldown = 1.0  # Minimum time between reps in seconds
        
        while is_workout_running and cap.isOpened():
            ret, frame = cap.read()
            if not ret:
                print("Failed to grab frame")
                break
                
            # Get image dimensions for MediaPipe
            h, w, _ = frame.shape
            
            # Flip the frame horizontally for a selfie-view display
            frame = cv2.flip(frame, 1)
            
            # To improve performance, mark the image as not writeable to pass by reference
            frame.flags.writeable = False
            
            # Convert frame to RGB
            frame_rgb = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
            
            # Process the frame with MediaPipe Pose
            results = pose.process(frame_rgb)
            
            # Set the frame to writeable again to modify it
            frame.flags.writeable = True
            
            # Convert back to BGR for display
            frame = cv2.cvtColor(frame_rgb, cv2.COLOR_RGB2BGR)
            
            if results.pose_landmarks:
                # Draw pose landmarks on the frame
                mp_drawing.draw_landmarks(
                    frame, 
                    results.pose_landmarks,
                    mp_pose.POSE_CONNECTIONS,
                    mp_drawing.DrawingSpec(color=(245,117,66), thickness=2, circle_radius=2),
                    mp_drawing.DrawingSpec(color=(245,66,230), thickness=2, circle_radius=2)
                )
                
                # Get landmarks
                landmarks = results.pose_landmarks.landmark
                
                # Detect exercise with cooldown
                current_time = time.time()
                is_rep = detect_exercise(landmarks, workout_type)
                
                # Count reps with cooldown
                if is_rep and stage == None and (current_time - last_rep_time) > rep_cooldown:
                    stage = "down"
                elif not is_rep and stage == "down" and (current_time - last_rep_time) > rep_cooldown:
                    stage = "up"
                    rep_count += 1
                    last_rep_time = current_time
                    
                # Calculate duration and calories
                duration = int(time.time() - start_time)
                calories = int(duration * 0.1)  # Rough estimate: 0.1 calories per second
                
                # Update workout data
                workout_data = {
                    "repCount": rep_count,
                    "duration": duration,
                    "calories": calories,
                    "stage": stage
                }
                
                # Draw rep count and other info on frame
                cv2.putText(frame, f'Reps: {rep_count}', (10,30), 
                           cv2.FONT_HERSHEY_SIMPLEX, 1, (255,255,255), 2, cv2.LINE_AA)
                cv2.putText(frame, f'Duration: {duration}s', (10,70), 
                           cv2.FONT_HERSHEY_SIMPLEX, 1, (255,255,255), 2, cv2.LINE_AA)
                cv2.putText(frame, f'Calories: {calories}', (10,110), 
                           cv2.FONT_HERSHEY_SIMPLEX, 1, (255,255,255), 2, cv2.LINE_AA)
                cv2.putText(frame, f'Stage: {stage}', (10,150), 
                           cv2.FONT_HERSHEY_SIMPLEX, 1, (255,255,255), 2, cv2.LINE_AA)
                
                # Draw exercise-specific guidance
                if workout_type == "squat":
                    cv2.putText(frame, 'Keep knees behind toes', (10,190), 
                               cv2.FONT_HERSHEY_SIMPLEX, 0.7, (255,255,255), 2, cv2.LINE_AA)
                elif workout_type == "bicep_curl":
                    cv2.putText(frame, 'Keep elbows close to body', (10,190), 
                               cv2.FONT_HERSHEY_SIMPLEX, 0.7, (255,255,255), 2, cv2.LINE_AA)
                elif workout_type == "pushup":
                    cv2.putText(frame, 'Keep body straight', (10,190), 
                               cv2.FONT_HERSHEY_SIMPLEX, 0.7, (255,255,255), 2, cv2.LINE_AA)
            
            # Display the frame
            cv2.imshow('Workout Monitoring', frame)
            
            # Break loop on 'q' press
            if cv2.waitKey(10) & 0xFF == ord('q'):
                break
        
        # Release resources
        cap.release()
        cv2.destroyAllWindows()
        print("Workout monitoring stopped")

@app.route('/')
def index():
    return render_template('startworkout.html')

@app.route('/api/workout/start', methods=['POST'])
def start_workout():
    global is_workout_running, workout_thread, workout_data
    
    # Check if a workout is already running
    if is_workout_running:
        return jsonify({"status": "error", "message": "Workout already in progress"})
    
    # Reset workout data
    workout_data = {
        "repCount": 0,
        "duration": 0,
        "calories": 0,
        "stage": None
    }
    
    # Get workout type from request
    data = request.json
    workout_type = data.get('workoutType', 'squat')
    
    # Start workout monitoring in a separate thread
    is_workout_running = True
    workout_thread = threading.Thread(target=workout_monitoring, args=(workout_type,))
    workout_thread.daemon = True
    workout_thread.start()
    
    return jsonify({"status": "success", "message": f"{workout_type.capitalize()} workout started"})

@app.route('/api/workout/data', methods=['GET'])
def get_workout_data():
    return jsonify(workout_data)

@app.route('/api/workout/end', methods=['POST'])
def end_workout():
    global is_workout_running, workout_data
    
    # Check if a workout is running
    if not is_workout_running:
        return jsonify({"status": "error", "message": "No workout in progress"})
    
    # Stop workout monitoring
    is_workout_running = False
    
    # Wait for thread to finish
    if workout_thread:
        workout_thread.join(timeout=1)
    
    # Return final workout data
    return jsonify({
        "status": "success",
        "workout_data": workout_data
    })

@app.route('/bone_density.html')
def bone_density():
    return render_template('bone_density.html')

@app.route("/predict", methods=["POST"])
def predict_bone_density():
    try:
        print("Received bone density prediction request")
        if df is None:
            print("Bone density dataset not available")
            return jsonify({"error": "Bone density prediction service is not available"}), 503
            
        data = request.get_json()
        print(f"Request data: {data}")
        if not data:
            print("No data received in request")
            return jsonify({"error": "No data provided"}), 400
            
        age = float(data.get("age"))
        gender = data.get("gender")
        
        print(f"Received prediction request - Age: {age}, Gender: {gender}")

        if not gender or gender not in ['male', 'female']:
            print(f"Invalid gender: {gender}")
            return jsonify({"error": "Invalid gender"}), 400

        if age < 1 or age > 120:
            print(f"Invalid age: {age}")
            return jsonify({"error": "Age must be between 1 and 120"}), 400

        if gender == 'male' and model_male:
            predicted = model_male.predict([[age]])[0]
        elif gender == 'female' and model_female:
            predicted = model_female.predict([[age]])[0]
        else:
            print(f"Model not available for gender: {gender}")
            return jsonify({"error": "Prediction model not available"}), 503

        predicted = np.clip(predicted, min_bmd, max_bmd)
        score = ((predicted - min_bmd) / (max_bmd - min_bmd)) * 100
        score = np.clip(score, 0, 100)

        print(f"Prediction successful - BMD: {predicted}, Score: {score}")
        return jsonify({
            "predicted_bmd": float(predicted),
            "bmd_score": float(score)
        })
    except Exception as e:
        print(f"Error in bone density prediction: {str(e)}")
        return jsonify({"error": "Internal server error"}), 500

# Food database with nutrition information per 100g
FOOD_DATABASE = {
    1: {"name": "Chicken Breast", "calories": 165, "protein": 31, "carbs": 0, "fats": 3.6, "unit": "g"},
    2: {"name": "Brown Rice", "calories": 216, "protein": 4.5, "carbs": 45, "fats": 1.8, "unit": "g"},
    3: {"name": "Salmon", "calories": 208, "protein": 22, "carbs": 0, "fats": 13, "unit": "g"},
    4: {"name": "Sweet Potato", "calories": 86, "protein": 1.6, "carbs": 20, "fats": 0.1, "unit": "g"},
    5: {"name": "Broccoli", "calories": 34, "protein": 2.8, "carbs": 6.6, "fats": 0.4, "unit": "g"},
    6: {"name": "Greek Yogurt", "calories": 59, "protein": 10, "carbs": 3.6, "fats": 0.4, "unit": "g"},
    7: {"name": "Almonds", "calories": 579, "protein": 21, "carbs": 22, "fats": 50, "unit": "g"},
    8: {"name": "Banana", "calories": 105, "protein": 1.3, "carbs": 27, "fats": 0.3, "unit": "g"},
    9: {"name": "Egg", "calories": 155, "protein": 13, "carbs": 1.1, "fats": 11, "unit": "g"},
    10: {"name": "Oatmeal", "calories": 307, "protein": 13, "carbs": 55, "fats": 5, "unit": "g"},
    11:{"name":"Pizza","calories":266,"protein":11,"carbs":33,"fats":10,"unit":"g"},
    12:{"name":"Hamburger","calories":354,"protein":20,"carbs":29,"fats":17,"unit":"g"},
    13:{"name":"Salad","calories":20,"protein":1,"carbs":4,"fats":0.2,"unit":"g"},
    14:{"name":"Tomato","calories":18,"protein":1,"carbs":4,"fats":0.2,"unit":"g"},
    15:{"name":"Onion","calories":41,"protein":1,"carbs":9,"fats":0.1,"unit":"g"},
    16:{"name":"Carrot","calories":41,"protein":0.9,"carbs":9.6,"fats":0.2,"unit":"g"},
    17:{"name":"Potato","calories":77,"protein":2,"carbs":17,"fats":0.1,"unit":"g"},
    18:{"name":"Fish","calories":206,"protein":22,"carbs":0,"fats":12,"unit":"g"},
    19:{"name":"Bread","calories":265,"protein":9,"carbs":49,"fats":3.2,"unit":"g"},
    20:{"name":"Orange","calories":62,"protein":1.2,"carbs":15.4,"fats":0.2,"unit":"g"},
    21:{"name":"Pasta","calories":100,"protein":5,"carbs":20,"fats":1,"unit":"g"},
    22:{"name":"Burger","calories":250,"protein":15,"carbs":25,"fats":15,"unit":"g"},
    23:{"name":"Pizza","calories":266,"protein":11,"carbs":33,"fats":10,"unit":"g"},
    24:{"name":"Hamburger","calories":354,"protein":20,"carbs":29,"fats":17,"unit":"g"},
    25:{"name":"Salad","calories":20,"protein":1,"carbs":4,"fats":0.2,"unit":"g"},
    26:{"name":"Tomato","calories":18,"protein":1,"carbs":4,"fats":0.2,"unit":"g"},
    27:{"name":"Onion","calories":41,"protein":1,"carbs":9,"fats":0.1,"unit":"g"},
    28:{"name":"Carrot","calories":41,"protein":0.9,"carbs":9.6,"fats":0.2,"unit":"g"},
    29:{"name":"Potato","calories":77,"protein":2,"carbs":17,"fats":0.1,"unit":"g"},
    30:{"name":"Fish","calories":206,"protein":22,"carbs":0,"fats":12,"unit":"g"},
    31:{"name":"Bread","calories":265,"protein":9,"carbs":49,"fats":3.2,"unit":"g"},
    32:{"name":"Orange","calories":62,"protein":1.2,"carbs":15.4,"fats":0.2,"unit":"g"},
    33:{"name":"Pasta","calories":100,"protein":5,"carbs":20,"fats":1,"unit":"g"},
    34:{"name":"Burger","calories":250,"protein":15,"carbs":25,"fats":15,"unit":"g"},
    35:{"name":"Pizza","calories":266,"protein":11,"carbs":33,"fats":10,"unit":"g"},
    36:{"name":"Hamburger","calories":354,"protein":20,"carbs":29,"fats":17,"unit":"g"},
    37:{"name":"Salad","calories":20,"protein":1,"carbs":4,"fats":0.2,"unit":"g"},
    38:{"name":"Tomato","calories":18,"protein":1,"carbs":4,"fats":0.2,"unit":"g"},
    39:{"name":"Rice","calories":130,"protein":2.7,"carbs":28,"fats":0.3,"unit":"g"},
    40:{"name":"Chicken","calories":165,"protein":31,"carbs":0,"fats":3.6,"unit":"g"},
    41:{"name":"Fish","calories":206,"protein":22,"carbs":0,"fats":12,"unit":"g"},
    42:{"name":"Bread","calories":265,"protein":9,"carbs":49,"fats":3.2,"unit":"g"},
    43:{"name":"Orange","calories":62,"protein":1.2,"carbs":15.4,"fats":0.2,"unit":"g"},
    44:{"name":"Beef","calories":184,"protein":26,"carbs":0,"fats":11,"unit":"g"},
    45:{"name":"Pasta","calories":100,"protein":5,"carbs":20,"fats":1,"unit":"g"},
    46:{"name":"Burger","calories":250,"protein":15,"carbs":25,"fats":15,"unit":"g"},
    47:{"name":"Pizza","calories":266,"protein":11,"carbs":33,"fats":10,"unit":"g"},
    48:{"name":"Hamburger","calories":354,"protein":20,"carbs":29,"fats":17,"unit":"g"},
    49:{"name":"Salad","calories":20,"protein":1,"carbs":4,"fats":0.2,"unit":"g"},
    
    
    
    
    
}

# Food nutrition information for detected items
FOOD_NUTRITION = {
    'apple': {'calories': 95, 'protein': 0.5, 'carbs': 25, 'fat': 0.3},
    'banana': {'calories': 105, 'protein': 1.3, 'carbs': 27, 'fat': 0.3},
    'pizza': {'calories': 266, 'protein': 11, 'carbs': 33, 'fat': 10},
    'hamburger': {'calories': 354, 'protein': 20, 'carbs': 29, 'fat': 17},
    'salad': {'calories': 20, 'protein': 1, 'carbs': 4, 'fat': 0.2},
    'rice': {'calories': 130, 'protein': 2.7, 'carbs': 28, 'fat': 0.3},
    'chicken': {'calories': 165, 'protein': 31, 'carbs': 0, 'fat': 3.6},
    'fish': {'calories': 206, 'protein': 22, 'carbs': 0, 'fat': 12},
    'bread': {'calories': 265, 'protein': 9, 'carbs': 49, 'fat': 3.2},
    'egg': {'calories': 72, 'protein': 6.3, 'carbs': 0.6, 'fat': 4.8},
    'orange': {'calories': 62, 'protein': 1.2, 'carbs': 15.4, 'fat': 0.2},
    'carrot': {'calories': 41, 'protein': 0.9, 'carbs': 9.6, 'fat': 0.2},
    'broccoli': {'calories': 34, 'protein': 2.8, 'carbs': 6.6, 'fat': 0.4},
    'tomato': {'calories': 22, 'protein': 1.1, 'carbs': 4.8, 'fat': 0.2},
    'potato': {'calories': 77, 'protein': 2, 'carbs': 17, 'fat': 0.1}
}

# Unit conversion factors (to grams)
UNIT_CONVERSIONS = {
    "g": 1,
    "oz": 28.35,
    "cup": 236.59,
    "tbsp": 14.79,
    "tsp": 4.93
}

@app.route('/api/food/search')
def search_food():
    query = request.args.get('query', '').lower()
    if not query:
        return jsonify([])
    
    # Search for food items that match the query
    results = []
    for food_id, food in FOOD_DATABASE.items():
        if query in food['name'].lower():
            results.append({
                'id': food_id,
                'name': food['name']
            })
    
    return jsonify(results[:10])  # Limit to 10 results

@app.route('/api/food/calculate-nutrients', methods=['POST'])
def calculate_nutrients():
    try:
        data = request.get_json()
        print("Received data:", data)  # Debug log
        
        food_items = data.get('foodItems', [])
        print("Food items:", food_items)  # Debug log
        
        if not food_items:
            return jsonify({'error': 'No food items provided'}), 400
        
        total_calories = 0
        total_protein = 0
        total_carbs = 0
        total_fats = 0
        
        calculated_items = []
        
        for item in food_items:
            print("Processing item:", item)  # Debug log
            
            food_id = int(item.get('id'))
            quantity = float(item.get('quantity', 0))
            unit = item.get('unit', 'g')
            
            print(f"Food ID: {food_id}, Quantity: {quantity}, Unit: {unit}")  # Debug log
            
            if food_id not in FOOD_DATABASE:
                print(f"Food ID {food_id} not found in database")  # Debug log
                continue
            
            food = FOOD_DATABASE[food_id]
            print("Found food:", food)  # Debug log
            
            # Convert quantity to grams
            quantity_in_grams = quantity * UNIT_CONVERSIONS.get(unit, 1)
            print(f"Quantity in grams: {quantity_in_grams}")  # Debug log
            
            # Calculate nutrients based on quantity (values are per 100g)
            scale_factor = quantity_in_grams / 100.0
            calories = food['calories'] * scale_factor
            protein = food['protein'] * scale_factor
            carbs = food['carbs'] * scale_factor
            fats = food['fats'] * scale_factor
            
            print(f"Calculated values - Calories: {calories}, Protein: {protein}, Carbs: {carbs}, Fats: {fats}")  # Debug log
            
            # Add to totals
            total_calories += calories
            total_protein += protein
            total_carbs += carbs
            total_fats += fats
            
            # Add to calculated items
            calculated_items.append({
                'name': food['name'],
                'quantity': quantity,
                'unit': unit,
                'calories': calories,
                'protein': protein,
                'carbs': carbs,
                'fats': fats
            })
        
        result = {
            'total': {
                'calories': total_calories,
                'protein': total_protein,
                'carbs': total_carbs,
                'fats': total_fats
            },
            'items': calculated_items
        }
        
        print("Final result:", result)  # Debug log
        return jsonify(result)
        
    except Exception as e:
        print(f"Error calculating nutrients: {str(e)}")  # Debug log
        return jsonify({'error': str(e)}), 500

# Load YOLOv5 model
try:
    print("Loading YOLOv5 model...")
    model = torch.hub.load('ultralytics/yolov5', 'yolov5s', pretrained=True)
    print("YOLOv5 model loaded successfully")
except Exception as e:
    print(f"Error loading YOLOv5 model: {str(e)}")
    model = None

@app.route('/api/detect-food', methods=['POST'])
def detect_food():
    try:
        if 'image' not in request.files:
            return jsonify({'error': 'No image provided'}), 400
            
        # Get the image file
        image_file = request.files['image']
        
        # Read the image
        img = Image.open(image_file)
        
        # Convert to RGB if needed
        if img.mode != 'RGB':
            img = img.convert('RGB')
            
        # Perform object detection
        if model is not None:
            results = model(img)
            
            # Get detections
            detections = []
            for *box, conf, cls in results.xyxy[0]:
                label = model.names[int(cls)]
                confidence = float(conf)
                
                # Lower the confidence threshold to detect more items
                if confidence > 0.3:
                    # Try to match the label with our food database
                    matched_label = None
                    
                    # Check for rice-related items
                    if 'rice' in label.lower():
                        matched_label = 'rice'
                    # Check for other food items
                    elif label.lower() in FOOD_NUTRITION:
                        matched_label = label.lower()
                    # Try to match with similar names
                    else:
                        for food_name in FOOD_NUTRITION.keys():
                            if food_name in label.lower() or label.lower() in food_name:
                                matched_label = food_name
                                break
                    
                    if matched_label:
                        detections.append({
                            'label': matched_label,
                            'confidence': confidence,
                            'nutrition': FOOD_NUTRITION[matched_label]
                        })
            
            # Draw bounding boxes on the image
            img_np = np.array(img)
            for detection in detections:
                # Get the box coordinates
                box = results.xyxy[0][detections.index(detection)][:4].cpu().numpy()
                x1, y1, x2, y2 = map(int, box)
                
                # Draw rectangle
                cv2.rectangle(img_np, (x1, y1), (x2, y2), (0, 255, 0), 2)
                
                # Add label and confidence
                label = f"{detection['label']} ({detection['confidence']:.2f})"
                cv2.putText(img_np, label, (x1, y1-10), cv2.FONT_HERSHEY_SIMPLEX, 0.5, (0, 255, 0), 2)
            
            # Convert annotated image to base64
            _, buffer = cv2.imencode('.jpg', img_np)
            img_base64 = base64.b64encode(buffer).decode('utf-8')
            
            return jsonify({
                'image': img_base64,
                'detections': detections
            })
        else:
            return jsonify({'error': 'Model not loaded'}), 500
            
    except Exception as e:
        print(f"Error in food detection: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/get-nutrition', methods=['POST'])
def get_nutrition():
    try:
        data = request.get_json()
        food_name = data.get('foodName', '').lower()
        quantity = float(data.get('quantity', 100))  # Default to 100g if not specified
        
        # Find the food in our database
        food_data = None
        for food_key in FOOD_NUTRITION:
            if food_name in food_key or food_key in food_name:
                food_data = FOOD_NUTRITION[food_key]
                break
        
        if not food_data:
            return jsonify({'error': 'Food not found in database'}), 404
        
        # Calculate nutrition based on quantity (assuming values are per 100g)
        scale_factor = quantity / 100.0
        nutrition_info = {
            'calories': round(food_data['calories'] * scale_factor, 1),
            'protein': round(food_data['protein'] * scale_factor, 1),
            'carbs': round(food_data['carbs'] * scale_factor, 1),
            'fat': round(food_data['fat'] * scale_factor, 1)
        }
        
        return jsonify(nutrition_info)
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/export-meal-report', methods=['POST'])
def export_meal_report():
    try:
        data = request.get_json()
        if not data or 'items' not in data:
            return jsonify({'error': 'No meal data provided'}), 400

        # Create a new workbook and select the active sheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Meal Report"

        # Add title
        ws['A1'] = "Meal Report"
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:G1')
        ws['A1'].alignment = Alignment(horizontal='center')

        # Add date
        ws['A2'] = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws.merge_cells('A2:G2')
        ws['A2'].alignment = Alignment(horizontal='center')

        # Add headers
        headers = ['Meal Name', 'Type', 'Time', 'Calories', 'Protein (g)', 'Carbs (g)', 'Fats (g)', 'Notes']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')

        # Add meal items
        row = 5
        total_calories = 0
        total_protein = 0
        total_carbs = 0
        total_fats = 0

        for item in data['items']:
            ws.cell(row=row, column=1).value = item['name']
            ws.cell(row=row, column=2).value = item['type'].capitalize()
            ws.cell(row=row, column=3).value = item['time']
            ws.cell(row=row, column=4).value = round(item['calories'], 1)
            ws.cell(row=row, column=5).value = round(item['protein'], 1)
            ws.cell(row=row, column=6).value = round(item['carbs'], 1)
            ws.cell(row=row, column=7).value = round(item['fats'], 1)
            ws.cell(row=row, column=8).value = item['notes']

            total_calories += item['calories']
            total_protein += item['protein']
            total_carbs += item['carbs']
            total_fats += item['fats']
            row += 1

        # Add totals row
        ws.cell(row=row, column=1).value = "TOTAL"
        ws.cell(row=row, column=1).font = Font(bold=True)
        ws.cell(row=row, column=4).value = round(total_calories, 1)
        ws.cell(row=row, column=5).value = round(total_protein, 1)
        ws.cell(row=row, column=6).value = round(total_carbs, 1)
        ws.cell(row=row, column=7).value = round(total_fats, 1)

        # Style the totals row
        for col in range(1, 9):
            cell = ws.cell(row=row, column=col)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")

        # Adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width

        # Save the workbook to a BytesIO object
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        # Generate filename with timestamp
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'meal_report_{timestamp}.xlsx'

        return send_file(
            excel_file,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        print(f"Error exporting meal report: {str(e)}")
        return jsonify({'error': str(e)}), 500

if __name__ == '__main__':
    # Print instructions
    print("Starting Workout Tracker Application")
    print("API endpoints:")
    print("  - POST /api/workout/start - Start a workout")
    print("  - GET /api/workout/data - Get current workout data")
    print("  - POST /api/workout/end - End the current workout")
    print("  - POST /api/detect-food - Analyze food image")
    
    # Start Flask app with logging disabled
    import sys
    cli = sys.modules['flask.cli']
    cli.show_server_banner = lambda *x: None  # Hide Flask server banner
    app.run(debug=False, port=5000, use_reloader=False)  # Set debug=False to reduce logging